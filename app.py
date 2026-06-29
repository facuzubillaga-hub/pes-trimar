import os, json, re, base64, sqlite3
from flask import Flask, request, jsonify, send_file, render_template, g, session, redirect, url_for
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import anthropic
import tempfile
from plan_cargas_parser import parse_plan_cargas
from gmail_helper import (
    get_flow, get_credentials, save_credentials, get_gmail_service,
    get_pdf_attachments_from_message, fetch_new_emails,
    mark_as_read, get_message_sender,
    BUNGE_SENDERS, PE_SENDER_DOMAIN
)

app = Flask(__name__)
app.secret_key = os.environ.get("FLASK_SECRET", "trimar-secret-2026")

DATA_DIR = "/data" if os.path.exists("/data") else os.path.dirname(__file__)
DB_PATH = os.path.join(DATA_DIR, "permisos.db")

HEADERS_PE = [
    "Nº PE", "Fecha Oficialización", "Buque", "Bandera", "País Destino",
    "Mercadería", "Posición Arancelaria", "Cond. Venta",
    "Peso Bruto (kg)", "Toneladas", "FOB Total (USD)", "Flete (USD)",
    "Precio Unit. (USD/Tn)", "Precio Oficial (USD/Tn)",
    "DJVE", "Fecha Cierre Vta.", "Vto. Embarque", "Ant. Gan. (USD)", "Cotización"
]
FIELDS_PE = [
    "nro_pe", "fecha_oficializacion", "buque", "bandera", "pais_destino",
    "mercaderia", "posicion_arancelaria", "cond_venta",
    "peso_bruto_kg", "toneladas", "fob_total_usd", "flete_usd",
    "precio_unit_usd_tn", "precio_oficial_usd_tn",
    "djve", "fecha_cierre_vta", "vto_embarque", "ant_gan_usd", "cotizacion"
]
HEADERS_SOL = [
    "Buque", "Bandera", "Destino", "Producto", "Cantidad (Tn)",
    "DJVE", "Contrato", "Cond. Venta", "Precio FOB (USD/Tn)",
    "Importe Total (USD)", "Fecha Solicitud", "Fecha Vta.", "Booking", "Estado PE"
]
FIELDS_SOL = [
    "buque", "bandera", "destino", "producto", "cantidad_tn",
    "djve", "contrato", "cond_venta", "precio_fob_usd_tn",
    "importe_total_usd", "fecha_solicitud", "fecha_vta", "booking"
]
FIELDS_PC = [
    "semana", "buque", "booking", "producto", "consignee",
    "cantidad_contenedores", "toneladas", "pod", "etd", "fecha_carga",
    "contrato", "envase", "linea"
]

# ---------- DB ----------

def get_db():
    db = getattr(g, '_database', None)
    if db is None:
        db = g._database = sqlite3.connect(DB_PATH)
        db.row_factory = sqlite3.Row
    return db

@app.teardown_appcontext
def close_db(exc):
    db = getattr(g, '_database', None)
    if db is not None:
        db.close()

def init_db():
    with app.app_context():
        db = get_db()
        db.execute(f"""CREATE TABLE IF NOT EXISTS permisos (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            {', '.join(f'{f} TEXT' for f in FIELDS_PE)},
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        db.execute(f"""CREATE TABLE IF NOT EXISTS solicitudes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            {', '.join(f'{f} TEXT' for f in FIELDS_SOL)},
            gmail_msg_id TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        db.execute(f"""CREATE TABLE IF NOT EXISTS plan_cargas (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            {', '.join(f'{f} TEXT' for f in FIELDS_PC)},
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
        db.commit()
        count = db.execute("SELECT COUNT(*) FROM permisos").fetchone()[0]
        if count == 0:
            seed = [
                ("26040EC01000700V","13/05/2026","EVER FEAT","LIBERIA","MALASIA","Aceite de Girasol (a granel)","1512.11.10.919G","FCA",264000,264,371712.00,None,1408.00,1288,"26001DJVE001338V","21/04/2026","27/06/2026",1700.16,1384.00),
                ("26040EC01000719X","15/05/2026","MERCOSUL ITAJAI","BRASIL","MALASIA","Aceite de Girasol (a granel)","1512.11.10.919G","CFR",198000,198,255024.00,4356.00,1288.00,1288,"26001DJVE001392V","24/04/2026","29/06/2026",1275.12,1391.00),
                ("26040EC01000720A","15/05/2026","MERCOSUL SANTOS","BRASIL","MALASIA","Aceite de Girasol (a granel)","1512.11.10.919G","CFR",198000,198,255024.00,4356.00,1288.00,1288,"26001DJVE001392V","24/04/2026","29/06/2026",1275.12,1391.00),
                ("26040EC01000744G","20/05/2026","MSC CHLOE","PORTUGAL","JAPÓN","Aceite de Girasol Alto Oleico (a granel)","1512.11.10.911P","CFR",315000,315,393368.85,11721.15,1248.79,1286,"26001DJVE001694D","18/05/2026","04/07/2026",2025.45,1398.00),
                ("26040EC01000745H","20/05/2026","MSC CHLOE","PORTUGAL","JAPÓN","Aceite de Girasol Alto Oleico (a granel)","1512.11.10.911P","CFR",315000,315,405090.00,11721.15,1286.00,1286,"26001DJVE001694D","18/05/2026","04/07/2026",2025.45,1398.00),
                ("26040EC01000764X","21/05/2026","CMA CGM RODOLPHE","SINGAPUR","MALASIA","Aceite de Girasol (a granel)","1512.11.10.919G","FCA",264000,264,371712.00,None,1408.00,1288,"26001DJVE001338V","21/04/2026","05/07/2026",1700.16,1397.00),
            ]
            ph = ','.join(['?']*len(FIELDS_PE))
            db.executemany(f"INSERT INTO permisos ({','.join(FIELDS_PE)}) VALUES ({ph})", seed)
            db.commit()

# ---------- Excel builder ----------

def _border():
    t = Side(style="thin")
    return Border(left=t, right=t, top=t, bottom=t)

def _hcell(cell, text):
    cell.value = text
    cell.font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    cell.fill = PatternFill("solid", start_color="1F4E79")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _border()

def build_excel(pe_rows, sol_rows, pc_rows):
    wb = Workbook()
    fills = [PatternFill("solid", start_color="DCE6F1"), PatternFill("solid", start_color="FFFFFF")]

    # ---- Hoja PEs ----
    ws1 = wb.active
    ws1.title = "Permisos de Exportación"
    for col, h in enumerate(HEADERS_PE, 1):
        _hcell(ws1.cell(row=1, column=col), h)
    ws1.row_dimensions[1].height = 35
    for r, row in enumerate(pe_rows, 2):
        for c, f in enumerate(FIELDS_PE, 1):
            val = row[f]
            try: val = float(val) if val and '.' in str(val) else (int(val) if val else None)
            except: pass
            cell = ws1.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border(); cell.fill = fills[r % 2]
    tr = len(pe_rows)+2
    for col in range(1, len(HEADERS_PE)+1):
        cell = ws1.cell(row=tr, column=col)
        cell.fill = PatternFill("solid", start_color="BDD7EE"); cell.border = _border()
        cell.font = Font(name="Arial", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws1.cell(row=tr, column=1).value = "TOTALES"
    for col in [9,10,11,12,18]:
        cl = get_column_letter(col)
        ws1.cell(row=tr, column=col).value = f"=SUM({cl}2:{cl}{tr-1})"
    for i, w in enumerate([20,18,20,12,14,35,20,14,16,12,16,14,18,20,22,18,16,16,12],1):
        ws1.column_dimensions[get_column_letter(i)].width = w
    ws1.freeze_panes = "A2"

    # ---- Hoja Solicitudes ----
    ws2 = wb.create_sheet("Solicitudes Bunge")
    HEADERS_SOL_FULL = ["Buque","Bandera","Destino","Producto","Cantidad (Tn)","DJVE","Contrato","Cond. Venta","Precio FOB (USD/Tn)","Importe Total (USD)","Fecha Solicitud","Fecha Vta.","Booking","Estado PE"]
    for col, h in enumerate(HEADERS_SOL_FULL, 1): _hcell(ws2.cell(row=1, column=col), h)
    ws2.row_dimensions[1].height = 35
    pe_djves = set(str(r["djve"] or "").strip() for r in pe_rows)
    pe_buques = set(str(r["buque"] or "").strip().upper() for r in pe_rows)
    gf = PatternFill("solid", start_color="C6EFCE")
    yf = PatternFill("solid", start_color="FFEB9C")
    for r, row in enumerate(sol_rows, 2):
        djve = str(row["djve"] or "").strip()
        buque = str(row["buque"] or "").strip().upper()
        tiene_pe = djve in pe_djves or buque in pe_buques
        rf = gf if tiene_pe else yf
        for c, f in enumerate(FIELDS_SOL, 1):
            val = row[f]
            try: val = float(val) if val and '.' in str(val) else (int(val) if val else None)
            except: pass
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border(); cell.fill = rf
        ec = ws2.cell(row=r, column=len(FIELDS_SOL)+1, value="✅ PE Generado" if tiene_pe else "⏳ Pendiente")
        ec.font = Font(name="Arial", bold=True, size=10)
        ec.alignment = Alignment(horizontal="center", vertical="center")
        ec.border = _border(); ec.fill = rf
    for i, w in enumerate([20,12,14,25,14,22,14,12,18,18,16,14,16,16],1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    ws2.freeze_panes = "A2"

    # ---- Hoja Plan de Cargas ----
    ws3 = wb.create_sheet("Plan de Cargas")
    HEADERS_PC_FULL = ["Semana","Buque","Booking","Producto","Consignee","Contenedores","Toneladas","Destino (POD)","ETD","Fecha Carga","Contrato","Envase","Línea","Solicitud","PE"]
    for col, h in enumerate(HEADERS_PC_FULL, 1): _hcell(ws3.cell(row=1, column=col), h)
    ws3.row_dimensions[1].height = 35
    sol_bookings = set(str(r["booking"] or "").strip() for r in sol_rows if r["booking"])
    sol_buques = set(str(r["buque"] or "").strip().upper() for r in sol_rows)
    for r, row in enumerate(pc_rows, 2):
        booking = str(row["booking"] or "").strip()
        buque = str(row["buque"] or "").strip().upper()
        tiene_sol = booking in sol_bookings or buque in sol_buques
        tiene_pe = buque in pe_buques
        if tiene_pe: rf = gf
        elif tiene_sol: rf = PatternFill("solid", start_color="DDEBF7")
        else: rf = yf
        for c, f in enumerate(FIELDS_PC, 1):
            cell = ws3.cell(row=r, column=c, value=str(row[f]) if row[f] else None)
            cell.font = Font(name="Arial", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _border(); cell.fill = rf
        ws3.cell(row=r, column=14, value="✅" if tiene_sol else "—").fill = rf
        ws3.cell(row=r, column=15, value="✅" if tiene_pe else "—").fill = rf
        for c in [14, 15]:
            ws3.cell(row=r, column=c).font = Font(name="Arial", size=10)
            ws3.cell(row=r, column=c).alignment = Alignment(horizontal="center", vertical="center")
            ws3.cell(row=r, column=c).border = _border()
    for i, w in enumerate([14,20,18,30,16,14,12,16,12,12,14,14,14,12,10],1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    ws3.freeze_panes = "A2"

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    wb.save(tmp.name)
    return tmp.name

# ---------- Claude ----------

def _call_claude(pdf_bytes, prompt):
    client = anthropic.Anthropic()
    pdf_b64 = base64.standard_b64encode(pdf_bytes).decode("utf-8")
    response = client.messages.create(
        model="claude-sonnet-4-6", max_tokens=1000,
        messages=[{"role": "user", "content": [
            {"type": "document", "source": {"type": "base64", "media_type": "application/pdf", "data": pdf_b64}},
            {"type": "text", "text": prompt}
        ]}]
    )
    text = re.sub(r'^```json\s*', '', response.content[0].text.strip())
    text = re.sub(r'\s*```$', '', text)
    return json.loads(text)

PE_PROMPT = """Extraé los siguientes campos de este Permiso de Exportación (PE) argentino y devolvé SOLO un JSON válido:
{"nro_pe":"","fecha_oficializacion":"DD/MM/AAAA","buque":"","bandera":"","pais_destino":"","mercaderia":"","posicion_arancelaria":"","cond_venta":"","peso_bruto_kg":0,"toneladas":0,"fob_total_usd":0.0,"flete_usd":null,"precio_unit_usd_tn":0.0,"precio_oficial_usd_tn":0,"djve":"","fecha_cierre_vta":"DD/MM/AAAA","vto_embarque":"DD/MM/AAAA","ant_gan_usd":0.0,"cotizacion":0.0}
Solo el JSON, nada más."""

SOL_PROMPT = """Extraé los siguientes campos de esta Solicitud de Permiso de Embarque de Bunge y devolvé SOLO un JSON válido:
{"buque":"","bandera":"","destino":"","producto":"","cantidad_tn":0.0,"djve":"","contrato":"","cond_venta":"","precio_fob_usd_tn":0.0,"importe_total_usd":0.0,"fecha_solicitud":"DD/MM/AAAA","fecha_vta":"DD/MM/AAAA","booking":null}
IMPORTANTE: cantidad_tn debe ser en TONELADAS (no kilogramos). Si ves un valor en kg como 198000, convertilo a toneladas: 198.
Solo el JSON, nada más."""

# ---------- Gmail OAuth ----------

def get_redirect_uri():
    base = os.environ.get("RAILWAY_PUBLIC_DOMAIN", "")
    if base:
        return f"https://{base}/oauth2callback"
    return "http://localhost:5000/oauth2callback"

@app.route("/gmail/auth")
def gmail_auth():
    client_id = os.environ.get("GOOGLE_CLIENT_ID")
    client_secret = os.environ.get("GOOGLE_CLIENT_SECRET")
    if not client_id or not client_secret:
        return jsonify({"error": "Faltan GOOGLE_CLIENT_ID y GOOGLE_CLIENT_SECRET"}), 500
    flow = get_flow(client_id, client_secret, get_redirect_uri())
    auth_url, state = flow.authorization_url(prompt="consent", access_type="offline")
    session["oauth_state"] = state
    return redirect(auth_url)

@app.route("/oauth2callback")
def oauth2callback():
    client_id = os.environ.get("GOOGLE_CLIENT_ID")
    client_secret = os.environ.get("GOOGLE_CLIENT_SECRET")
    flow = get_flow(client_id, client_secret, get_redirect_uri())
    flow.fetch_token(authorization_response=request.url)
    save_credentials(flow.credentials)
    return redirect("/?gmail=ok")

@app.route("/gmail/status")
def gmail_status():
    creds = get_credentials()
    return jsonify({"connected": creds is not None})

@app.route("/gmail/disconnect")
def gmail_disconnect():
    token_path = os.path.join(DATA_DIR, "gmail_token.json")
    if os.path.exists(token_path):
        os.remove(token_path)
    return jsonify({"ok": True})

# ---------- Gmail fetch ----------

@app.route("/gmail/fetch/<tipo>", methods=["POST"])
def gmail_fetch(tipo):
    service = get_gmail_service()
    if not service:
        return jsonify({"error": "Gmail no conectado"}), 401
    if tipo == "pe":
        query_senders = [f"@{PE_SENDER_DOMAIN}"]
    else:
        query_senders = BUNGE_SENDERS
    query = f"from:({'  OR '.join(query_senders)}) has:attachment filename:pdf is:unread"
    results = service.users().messages().list(userId="me", q=query, maxResults=20).execute()
    messages = results.get("messages", [])
    if not messages:
        return jsonify({"found": 0, "processed": [], "skipped": []})
    db = get_db()
    processed = []
    skipped = []
    for msg in messages:
        msg_id = msg["id"]
        sender, subject, date = get_message_sender(service, msg_id)
        attachments = get_pdf_attachments_from_message(service, msg_id)
        for filename, pdf_bytes in attachments:
            try:
                if tipo == "pe":
                    data = _call_claude(pdf_bytes, PE_PROMPT)
                    nro = data.get("nro_pe", "")
                    exists = db.execute("SELECT 1 FROM permisos WHERE nro_pe=?", (nro,)).fetchone()
                    if exists:
                        skipped.append({"filename": filename, "reason": f"PE {nro} ya existe"})
                        continue
                    ph = ','.join(['?']*len(FIELDS_PE))
                    vals = [str(data.get(f,'')) if data.get(f) is not None else None for f in FIELDS_PE]
                    db.execute(f"INSERT INTO permisos ({','.join(FIELDS_PE)}) VALUES ({ph})", vals)
                    db.commit()
                    processed.append({"filename": filename, "id": nro, "sender": sender, "subject": subject})
                else:
                    data = _call_claude(pdf_bytes, SOL_PROMPT)
                    exists = db.execute("SELECT 1 FROM solicitudes WHERE gmail_msg_id=?", (msg_id,)).fetchone()
                    if exists:
                        skipped.append({"filename": filename, "reason": "Ya procesado"})
                        continue
                    ph = ','.join(['?']*(len(FIELDS_SOL)+1))
                    vals = [str(data.get(f,'')) if data.get(f) is not None else None for f in FIELDS_SOL]
                    vals.append(msg_id)
                    db.execute(f"INSERT INTO solicitudes ({','.join(FIELDS_SOL)}, gmail_msg_id) VALUES ({ph})", vals)
                    db.commit()
                    processed.append({"filename": filename, "id": data.get("buque",""), "sender": sender, "subject": subject})
                mark_as_read(service, msg_id)
            except Exception as e:
                skipped.append({"filename": filename, "reason": str(e)})
    return jsonify({"found": len(messages), "processed": processed, "skipped": skipped})

# ---------- Routes PEs ----------

@app.route("/")
def index():
    return render_template("index.html")

@app.route("/upload", methods=["POST"])
def upload():
    if "pdf" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400
    pdf_bytes = request.files["pdf"].read()
    try:
        data = _call_claude(pdf_bytes, PE_PROMPT)
    except Exception as e:
        return jsonify({"error": f"Error extrayendo datos: {str(e)}"}), 500
    db = get_db()
    exists = db.execute("SELECT 1 FROM permisos WHERE nro_pe=?", (data.get("nro_pe",""),)).fetchone()
    data["already_exists"] = exists is not None
    djve = str(data.get("djve") or "").strip()
    toneladas = str(data.get("toneladas") or "").strip()
    sol = None
    if djve and toneladas:
        sol = db.execute("SELECT booking, buque, fecha_solicitud FROM solicitudes WHERE djve=? AND ROUND(CAST(cantidad_tn AS REAL))=ROUND(CAST(? AS REAL))", (djve, toneladas)).fetchone()
    if not sol and djve:
        sol = db.execute("SELECT booking, buque, fecha_solicitud FROM solicitudes WHERE djve=?", (djve,)).fetchone()
    data["solicitud_match"] = dict(sol) if sol else None
    return jsonify(data)

@app.route("/confirm", methods=["POST"])
def confirm():
    pe = request.json
    nro = pe.get("nro_pe", "")
    db = get_db()
    if db.execute("SELECT 1 FROM permisos WHERE nro_pe=?", (nro,)).fetchone():
        return jsonify({"error": f"El PE {nro} ya existe"}), 409
    pe.pop("already_exists", None)
    pe.pop("solicitud_match", None)
    ph = ','.join(['?']*len(FIELDS_PE))
    vals = [str(pe.get(f,'')) if pe.get(f) is not None else None for f in FIELDS_PE]
    db.execute(f"INSERT INTO permisos ({','.join(FIELDS_PE)}) VALUES ({ph})", vals)
    db.commit()
    return jsonify({"ok": True, "nro_pe": nro})

@app.route("/list")
def list_pes():
    db = get_db()
    rows = db.execute("SELECT nro_pe, buque, pais_destino, fecha_oficializacion, djve, toneladas FROM permisos ORDER BY fecha_oficializacion DESC, nro_pe DESC").fetchall()
    result = []
    for row in rows:
        d = dict(row)
        djve = str(d.get("djve") or "").strip()
        toneladas = str(d.get("toneladas") or "").strip()
        sol = None
        if djve and toneladas:
            sol = db.execute("SELECT booking FROM solicitudes WHERE djve=? AND ROUND(CAST(cantidad_tn AS REAL))=ROUND(CAST(? AS REAL))", (djve, toneladas)).fetchone()
        if not sol and djve:
            sol = db.execute("SELECT booking FROM solicitudes WHERE djve=?", (djve,)).fetchone()
        d["booking"] = sol["booking"] if sol and sol["booking"] else None
        result.append(d)
    return jsonify(result)

@app.route("/pe/delete/<nro>", methods=["DELETE"])
def pe_delete(nro):
    db = get_db()
    db.execute("DELETE FROM permisos WHERE nro_pe=?", (nro,))
    db.commit()
    return jsonify({"ok": True})

# ---------- Routes Solicitudes ----------

@app.route("/solicitud/upload", methods=["POST"])
def solicitud_upload():
    if "pdf" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400
    pdf_bytes = request.files["pdf"].read()
    try:
        data = _call_claude(pdf_bytes, SOL_PROMPT)
    except Exception as e:
        return jsonify({"error": f"Error extrayendo datos: {str(e)}"}), 500
    db = get_db()
    djve = data.get("djve","")
    buque = (data.get("buque") or "").upper()
    tiene_pe = bool(djve and db.execute("SELECT 1 FROM permisos WHERE djve=?", (djve,)).fetchone())
    if not tiene_pe and buque:
        tiene_pe = bool(db.execute("SELECT 1 FROM permisos WHERE UPPER(buque)=?", (buque,)).fetchone())
    data["tiene_pe"] = tiene_pe
    return jsonify(data)

@app.route("/solicitud/confirm", methods=["POST"])
def solicitud_confirm():
    sol = request.json
    sol.pop("tiene_pe", None)
    db = get_db()
    ph = ','.join(['?']*len(FIELDS_SOL))
    vals = [str(sol.get(f,'')) if sol.get(f) is not None else None for f in FIELDS_SOL]
    db.execute(f"INSERT INTO solicitudes ({','.join(FIELDS_SOL)}) VALUES ({ph})", vals)
    db.commit()
    return jsonify({"ok": True})

@app.route("/solicitud/list")
def list_solicitudes():
    db = get_db()
    rows = db.execute(f"SELECT {','.join(FIELDS_SOL)}, id FROM solicitudes ORDER BY fecha_solicitud DESC, id DESC").fetchall()
    pe_djves = set(str(r[0] or "").strip() for r in db.execute("SELECT djve FROM permisos WHERE djve IS NOT NULL").fetchall())
    pe_buques = set(str(r[0] or "").upper() for r in db.execute("SELECT buque FROM permisos WHERE buque IS NOT NULL").fetchall())
    result = []
    for row in rows:
        d = dict(row)
        djve_sol = str(d.get("djve") or "").strip()
        cant_sol = str(d.get("cantidad_tn") or "").strip()
        tiene_pe = False
        if djve_sol and cant_sol:
            tiene_pe = bool(db.execute("SELECT 1 FROM permisos WHERE djve=? AND ROUND(CAST(toneladas AS REAL))=ROUND(CAST(? AS REAL))", (djve_sol, cant_sol)).fetchone())
        d["tiene_pe"] = tiene_pe
        result.append(d)
    return jsonify(result)

@app.route("/solicitud/delete/<int:sol_id>", methods=["DELETE"])
def solicitud_delete(sol_id):
    db = get_db()
    db.execute("DELETE FROM solicitudes WHERE id=?", (sol_id,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/dashboard")
def dashboard():
    db = get_db()
    from datetime import datetime, timedelta

    total_pes = db.execute("SELECT COUNT(*) FROM permisos").fetchone()[0]
    total_sol = db.execute("SELECT COUNT(*) FROM solicitudes").fetchone()[0]
    total_plan = db.execute("SELECT COUNT(*) FROM plan_cargas").fetchone()[0]

    # Solicitudes sin PE
    sol_rows = db.execute(f"SELECT {','.join(FIELDS_SOL)} FROM solicitudes").fetchall()
    pe_djves = set(str(r[0] or "").strip() for r in db.execute("SELECT djve FROM permisos WHERE djve IS NOT NULL").fetchall())
    sol_sin_pe = 0
    for row in sol_rows:
        djve_sol = str(row["djve"] or "").strip()
        cant_sol = str(row["cantidad_tn"] or "").strip()
        tiene_pe = False
        if djve_sol and cant_sol:
            tiene_pe = bool(db.execute("SELECT 1 FROM permisos WHERE djve=? AND ROUND(CAST(toneladas AS REAL))=ROUND(CAST(? AS REAL))", (djve_sol, cant_sol)).fetchone())
        if not tiene_pe and djve_sol:
            tiene_pe = djve_sol in pe_djves
        if not tiene_pe:
            sol_sin_pe += 1

    # PEs por vencer en 7 días o vencidos
    today = datetime.today()
    pes_vencer = []
    for row in db.execute("SELECT nro_pe, buque, vto_embarque, pais_destino FROM permisos WHERE vto_embarque IS NOT NULL").fetchall():
        vto = row["vto_embarque"]
        if not vto:
            continue
        try:
            d = datetime.strptime(vto, "%d/%m/%Y")
            diff = (d - today).days
            if diff <= 7:
                pes_vencer.append({
                    "nro_pe": row["nro_pe"],
                    "buque": row["buque"],
                    "vto_embarque": vto,
                    "pais_destino": row["pais_destino"],
                    "dias": diff
                })
        except:
            pass
    pes_vencer.sort(key=lambda x: x["dias"])

    # Últimos 5 PEs
    ultimos_pes = [dict(r) for r in db.execute("SELECT nro_pe, buque, fecha_oficializacion, pais_destino FROM permisos ORDER BY fecha_oficializacion DESC, nro_pe DESC LIMIT 5").fetchall()]

    # Últimas 5 solicitudes
    ultimas_sol = [dict(r) for r in db.execute("SELECT buque, producto, cantidad_tn, djve, fecha_solicitud FROM solicitudes ORDER BY id DESC LIMIT 5").fetchall()]

    return jsonify({
        "total_pes": total_pes,
        "total_sol": total_sol,
        "total_plan": total_plan,
        "sol_sin_pe": sol_sin_pe,
        "pes_vencer": pes_vencer,
        "ultimos_pes": ultimos_pes,
        "ultimas_sol": ultimas_sol
    })


@app.route("/search")
def search():
    q = request.args.get("q", "").strip()
    if not q or len(q) < 2:
        return jsonify({"pes": [], "solicitudes": [], "plan": []})
    db = get_db()
    like = f"%{q.upper()}%"

    pes = [dict(r) for r in db.execute(f"""
        SELECT {','.join(FIELDS_PE)} FROM permisos
        WHERE UPPER(nro_pe) LIKE ? OR UPPER(buque) LIKE ? OR UPPER(djve) LIKE ?
        OR UPPER(pais_destino) LIKE ?
        ORDER BY fecha_oficializacion DESC LIMIT 20
    """, (like, like, like, like)).fetchall()]

    solicitudes = [dict(r) for r in db.execute(f"""
        SELECT {','.join(FIELDS_SOL)}, id FROM solicitudes
        WHERE UPPER(buque) LIKE ? OR UPPER(booking) LIKE ? OR UPPER(djve) LIKE ?
        OR UPPER(destino) LIKE ?
        ORDER BY fecha_solicitud DESC LIMIT 20
    """, (like, like, like, like)).fetchall()]

    plan = [dict(r) for r in db.execute(f"""
        SELECT {','.join(FIELDS_PC)}, id FROM plan_cargas
        WHERE UPPER(buque) LIKE ? OR UPPER(booking) LIKE ? OR UPPER(pod) LIKE ?
        ORDER BY etd DESC LIMIT 20
    """, (like, like, like)).fetchall()]

    return jsonify({"pes": pes, "solicitudes": solicitudes, "plan": plan})

# ---------- Routes Plan de Cargas ----------

@app.route("/plan/upload", methods=["POST"])
def plan_upload():
    if "excel" not in request.files:
        return jsonify({"error": "No se recibió archivo"}), 400
    file_bytes = request.files["excel"].read()
    try:
        rows = parse_plan_cargas(file_bytes)
    except Exception as e:
        return jsonify({"error": f"Error procesando Excel: {str(e)}"}), 500
    if not rows:
        return jsonify({"error": "No se encontraron datos en el archivo"}), 400
    return jsonify({"rows": rows, "count": len(rows)})

@app.route("/plan/confirm", methods=["POST"])
def plan_confirm():
    rows = request.json.get("rows", [])
    db = get_db()
    ph = ','.join(['?']*len(FIELDS_PC))
    inserted = 0
    skipped = 0
    for row in rows:
        # Check duplicate by booking + consignee + fecha_carga combination
        booking = row.get("booking")
        consignee = row.get("consignee")
        contrato = row.get("contrato")
        fecha_carga = row.get("fecha_carga")
        if booking and consignee and contrato:
            exists = db.execute(
                "SELECT 1 FROM plan_cargas WHERE booking=? AND consignee=? AND contrato=?",
                (booking, consignee, contrato)
            ).fetchone()
            if exists:
                skipped += 1
                continue
        elif booking and consignee and fecha_carga:
            exists = db.execute(
                "SELECT 1 FROM plan_cargas WHERE booking=? AND consignee=? AND fecha_carga=?",
                (booking, consignee, fecha_carga)
            ).fetchone()
            if exists:
                skipped += 1
                continue
        vals = [str(row.get(f,'')) if row.get(f) is not None else None for f in FIELDS_PC]
        db.execute(f"INSERT INTO plan_cargas ({','.join(FIELDS_PC)}) VALUES ({ph})", vals)
        inserted += 1
    db.commit()
    return jsonify({"ok": True, "inserted": inserted, "skipped": skipped})

@app.route("/plan/list")
def plan_list():
    db = get_db()
    rows = db.execute(f"SELECT {','.join(FIELDS_PC)}, id FROM plan_cargas ORDER BY etd DESC, id DESC").fetchall()
    sol_bookings = set(str(r[0] or "").strip() for r in db.execute("SELECT booking FROM solicitudes WHERE booking IS NOT NULL").fetchall())
    sol_buques = set(str(r[0] or "").strip().upper() for r in db.execute("SELECT buque FROM solicitudes WHERE buque IS NOT NULL").fetchall())
    pe_buques = set(str(r[0] or "").strip().upper() for r in db.execute("SELECT buque FROM permisos WHERE buque IS NOT NULL").fetchall())
    def fuzzy_match(name, name_set):
        if not name:
            return False
        if name in name_set:
            return True
        for s in name_set:
            if s and (s in name or name in s):
                return True
        return False

    # Build booking → tiene_pe map from solicitudes
    sol_rows_full = db.execute("SELECT booking, djve, cantidad_tn FROM solicitudes WHERE booking IS NOT NULL").fetchall()
    booking_tiene_pe = {}
    for sol in sol_rows_full:
        bk = str(sol["booking"] or "").strip()
        if not bk:
            continue
        djve = str(sol["djve"] or "").strip()
        cant = str(sol["cantidad_tn"] or "").strip()
        tiene_pe = False
        if djve and cant:
            tiene_pe = bool(db.execute(
                "SELECT 1 FROM permisos WHERE djve=? AND ROUND(CAST(toneladas AS REAL))=ROUND(CAST(? AS REAL))",
                (djve, cant)).fetchone())
        if not tiene_pe and djve:
            tiene_pe = bool(db.execute("SELECT 1 FROM permisos WHERE djve=?", (djve,)).fetchone())
        booking_tiene_pe[bk] = tiene_pe

    result = []
    for row in rows:
        d = dict(row)
        booking = str(d.get("booking") or "").strip()
        buque = str(d.get("buque") or "").strip().upper()
        # Primary: booking → solicitud → PE chain
        if booking and booking in booking_tiene_pe:
            d["tiene_solicitud"] = True
            d["tiene_pe"] = booking_tiene_pe[booking]
        else:
            d["tiene_solicitud"] = booking in sol_bookings or fuzzy_match(buque, sol_buques)
            d["tiene_pe"] = fuzzy_match(buque, pe_buques)
        result.append(d)
    return jsonify(result)

@app.route("/plan/delete/<int:row_id>", methods=["DELETE"])
def plan_delete(row_id):
    db = get_db()
    db.execute("DELETE FROM plan_cargas WHERE id=?", (row_id,))
    db.commit()
    return jsonify({"ok": True})


@app.route("/admin/fix_cantidades")
def fix_cantidades():
    """Fix cantidad_tn values that were stored in kg instead of tonnes."""
    db = get_db()
    rows = db.execute("SELECT id, cantidad_tn FROM solicitudes").fetchall()
    fixed = 0
    for row in rows:
        val = row["cantidad_tn"]
        if val:
            try:
                f = float(val)
                if f > 1000:  # Likely in kg
                    new_val = round(f / 1000, 3)
                    db.execute("UPDATE solicitudes SET cantidad_tn=? WHERE id=?", (str(new_val), row["id"]))
                    fixed += 1
            except:
                pass
    db.commit()
    return jsonify({"ok": True, "fixed": fixed})

# ---------- Download ----------

@app.route("/download")
def download():
    db = get_db()
    pe_rows = db.execute(f"SELECT {','.join(FIELDS_PE)} FROM permisos ORDER BY fecha_oficializacion, nro_pe").fetchall()
    sol_rows = db.execute(f"SELECT {','.join(FIELDS_SOL)} FROM solicitudes ORDER BY fecha_solicitud DESC").fetchall()
    pc_rows = db.execute(f"SELECT {','.join(FIELDS_PC)} FROM plan_cargas ORDER BY etd DESC").fetchall()
    path = build_excel(pe_rows, sol_rows, pc_rows)
    return send_file(path, as_attachment=True, download_name="permisos_exportacion.xlsx")

with app.app_context():
    init_db()

if __name__ == "__main__":
    print("✅ App corriendo en http://localhost:5000")
    app.run(debug=False, port=5000)
